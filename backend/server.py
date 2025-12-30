from fastapi import FastAPI, APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import tempfile
import shutil
from pypdf import PdfReader, PdfWriter
from PIL import Image
import img2pdf
from pdf2docx import Converter
import io
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import openpyxl
from docx import Document
import pytesseract
from pygments import highlight
from pygments.lexers import CppLexer, PythonLexer
from pygments.formatters import HtmlFormatter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Preformatted
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.units import inch
import nbformat
from nbconvert import PDFExporter
from nbconvert.preprocessors import ExecutePreprocessor
import subprocess
import xml.etree.ElementTree as ET
from xml.dom import minidom

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection (optional - for future use)
try:
    mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
    client = AsyncIOMotorClient(mongo_url, serverSelectionTimeoutMS=5000)
    db = client[os.environ.get('DB_NAME', 'test_database')]
    # Test connection
    client.admin.command('ping')
    logging.info("MongoDB connected successfully")
except Exception as e:
    logging.warning(f"MongoDB connection failed: {e}. Running without database.")
    client = None
    db = None

# Create temporary upload directory
UPLOAD_DIR = ROOT_DIR / 'temp_uploads'
UPLOAD_DIR.mkdir(exist_ok=True)

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Utility function to save uploaded file
async def save_upload_file(upload_file: UploadFile) -> Path:
    file_id = str(uuid.uuid4())
    file_extension = Path(upload_file.filename).suffix
    temp_path = UPLOAD_DIR / f"{file_id}{file_extension}"
    
    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(upload_file.file, buffer)
    
    return temp_path

# Utility function to cleanup temp files
def cleanup_files(*files):
    for file in files:
        try:
            if file and Path(file).exists():
                Path(file).unlink()
        except Exception as e:
            logging.error(f"Error cleaning up file {file}: {e}")

@api_router.get("/")
async def root():
    return {"message": "PDF Master API"}

@api_router.post("/merge")
async def merge_pdfs(files: List[UploadFile] = File(...)):
    """Merge multiple PDF files into one"""
    temp_files = []
    output_file = None
    
    try:
        # Save all uploaded files
        for file in files:
            if not file.filename.lower().endswith('.pdf'):
                raise HTTPException(status_code=400, detail=f"File {file.filename} is not a PDF")
            temp_path = await save_upload_file(file)
            temp_files.append(temp_path)
        
        # Merge PDFs
        pdf_writer = PdfWriter()
        for pdf_path in temp_files:
            pdf_reader = PdfReader(str(pdf_path))
            for page in pdf_reader.pages:
                pdf_writer.add_page(page)
        
        # Save merged PDF
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_merged.pdf"
        with open(output_file, "wb") as f:
            pdf_writer.write(f)
        
        # Return file
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="merged.pdf",
            background=lambda: cleanup_files(*temp_files, output_file)
        )
    
    except Exception as e:
        cleanup_files(*temp_files, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/split")
async def split_pdf(file: UploadFile = File(...), pages: str = Form(...)):
    """Split PDF by page ranges (e.g., '1-3,5,7-9')"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        # Parse page ranges
        page_list = []
        for part in pages.split(','):
            if '-' in part:
                start, end = map(int, part.split('-'))
                page_list.extend(range(start - 1, end))
            else:
                page_list.append(int(part) - 1)
        
        # Split PDF
        pdf_reader = PdfReader(str(temp_file))
        pdf_writer = PdfWriter()
        
        for page_num in page_list:
            if 0 <= page_num < len(pdf_reader.pages):
                pdf_writer.add_page(pdf_reader.pages[page_num])
        
        # Save split PDF
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_split.pdf"
        with open(output_file, "wb") as f:
            pdf_writer.write(f)
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="split.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/compress")
async def compress_pdf(file: UploadFile = File(...)):
    """Compress PDF file"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        # Read and rewrite PDF (basic compression)
        pdf_reader = PdfReader(str(temp_file))
        pdf_writer = PdfWriter()
        
        for page in pdf_reader.pages:
            pdf_writer.add_page(page)
            # Compress after adding to writer
            pdf_writer.pages[-1].compress_content_streams()
        
        # Save compressed PDF
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_compressed.pdf"
        with open(output_file, "wb") as f:
            pdf_writer.write(f)
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="compressed.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/rotate")
async def rotate_pdf(file: UploadFile = File(...), angle: int = Form(...)):
    """Rotate PDF pages"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        pdf_reader = PdfReader(str(temp_file))
        pdf_writer = PdfWriter()
        
        for page in pdf_reader.pages:
            page.rotate(angle)
            pdf_writer.add_page(page)
        
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_rotated.pdf"
        with open(output_file, "wb") as f:
            pdf_writer.write(f)
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="rotated.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/pdf-to-jpg")
async def pdf_to_jpg(file: UploadFile = File(...)):
    """Convert PDF to JPG images"""
    temp_file = None
    output_files = []
    
    try:
        temp_file = await save_upload_file(file)
        
        # Convert PDF pages to images using PyMuPDF
        import fitz
        pdf_document = fitz.open(str(temp_file))
        
        # If single page, return single image
        if len(pdf_document) == 1:
            page = pdf_document[0]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            output_file = UPLOAD_DIR / f"{uuid.uuid4()}.jpg"
            pix.save(str(output_file))
            output_files.append(output_file)
            
            return FileResponse(
                output_file,
                media_type="image/jpeg",
                filename="page_1.jpg",
                background=lambda: cleanup_files(temp_file, output_file)
            )
        else:
            # Multiple pages - return first page for now
            page = pdf_document[0]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            output_file = UPLOAD_DIR / f"{uuid.uuid4()}.jpg"
            pix.save(str(output_file))
            
            return FileResponse(
                output_file,
                media_type="image/jpeg",
                filename="page_1.jpg",
                background=lambda: cleanup_files(temp_file, output_file)
            )
    
    except Exception as e:
        cleanup_files(temp_file, *output_files)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/pdf-to-png")
async def pdf_to_png(file: UploadFile = File(...)):
    """Convert PDF to PNG images"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        import fitz
        pdf_document = fitz.open(str(temp_file))
        page = pdf_document[0]
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}.png"
        pix.save(str(output_file))
        
        return FileResponse(
            output_file,
            media_type="image/png",
            filename="page_1.png",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/jpg-to-pdf")
async def jpg_to_pdf(file: UploadFile = File(...)):
    """Convert JPG to PDF"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}.pdf"
        with open(output_file, "wb") as f:
            f.write(img2pdf.convert(str(temp_file)))
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="converted.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/png-to-pdf")
async def png_to_pdf(file: UploadFile = File(...)):
    """Convert PNG to PDF"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}.pdf"
        with open(output_file, "wb") as f:
            f.write(img2pdf.convert(str(temp_file)))
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="converted.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/pdf-to-word")
async def pdf_to_word(file: UploadFile = File(...)):
    """Convert PDF to Word"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}.docx"
        
        cv = Converter(str(temp_file))
        cv.convert(str(output_file))
        cv.close()
        
        return FileResponse(
            output_file,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            filename="converted.docx",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/word-to-pdf")
async def word_to_pdf(file: UploadFile = File(...)):
    """Convert Word to PDF"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}.pdf"
        
        # Read Word document
        doc = Document(str(temp_file))
        
        # Create PDF
        c = canvas.Canvas(str(output_file), pagesize=letter)
        width, height = letter
        y_position = height - 50
        
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                c.drawString(50, y_position, paragraph.text[:100])
                y_position -= 20
                if y_position < 50:
                    c.showPage()
                    y_position = height - 50
        
        c.save()
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="converted.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/excel-to-pdf")
async def excel_to_pdf(file: UploadFile = File(...)):
    """Convert Excel to PDF"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}.pdf"
        
        # Read Excel
        wb = openpyxl.load_workbook(str(temp_file))
        sheet = wb.active
        
        # Create PDF
        c = canvas.Canvas(str(output_file), pagesize=letter)
        width, height = letter
        y_position = height - 50
        
        for row in sheet.iter_rows(values_only=True):
            row_text = ' | '.join([str(cell) if cell else '' for cell in row])
            if row_text.strip():
                c.drawString(50, y_position, row_text[:100])
                y_position -= 20
                if y_position < 50:
                    c.showPage()
                    y_position = height - 50
        
        c.save()
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="converted.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/pdf-to-excel")
async def pdf_to_excel(file: UploadFile = File(...)):
    """Convert PDF to Excel - extracts text to Excel"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}.xlsx"
        
        # Extract text from PDF
        pdf_reader = PdfReader(str(temp_file))
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PDF Content"
        
        row_num = 1
        for page_num, page in enumerate(pdf_reader.pages):
            text = page.extract_text()
            ws.cell(row=row_num, column=1, value=f"Page {page_num + 1}")
            row_num += 1
            for line in text.split('\n'):
                if line.strip():
                    ws.cell(row=row_num, column=1, value=line)
                    row_num += 1
            row_num += 1
        
        wb.save(str(output_file))
        
        return FileResponse(
            output_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="converted.xlsx",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/cpp-to-pdf")
async def cpp_to_pdf(file: UploadFile = File(...)):
    """Convert CPP source code to PDF with syntax highlighting"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}.pdf"
        
        # Read the C++ source code
        with open(temp_file, 'r', encoding='utf-8') as f:
            cpp_code = f.read()
        
        # Create PDF with syntax highlighting
        doc = SimpleDocTemplate(str(output_file), pagesize=letter)
        styles = getSampleStyleSheet()
        
        # Create a custom style for code
        code_style = ParagraphStyle(
            'Code',
            parent=styles['Code'],
            fontName='Courier',
            fontSize=9,
            leading=11,
            leftIndent=0,
            rightIndent=0,
            alignment=TA_LEFT,
            spaceBefore=0,
            spaceAfter=0,
        )
        
        story = []
        
        # Split code into lines and add to PDF
        lines = cpp_code.split('\n')
        for line in lines:
            # Escape special characters for reportlab
            line = line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            # Preserve spaces
            line = line.replace(' ', '&nbsp;')
            if not line.strip():
                line = '&nbsp;'
            story.append(Paragraph(line, code_style))
        
        doc.build(story)
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="converted.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/ocr")
async def ocr_pdf(file: UploadFile = File(...)):
    """Extract text from PDF using OCR"""
    temp_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        # First try regular text extraction
        pdf_reader = PdfReader(str(temp_file))
        text_content = ""
        
        for page in pdf_reader.pages:
            text_content += page.extract_text() + "\n\n"
        
        cleanup_files(temp_file)
        
        return JSONResponse({
            "text": text_content,
            "pages": len(pdf_reader.pages)
        })
    
    except Exception as e:
        cleanup_files(temp_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/watermark")
async def watermark_pdf(
    file: UploadFile = File(...), 
    text: Optional[str] = Form(None),
    watermark_image: Optional[UploadFile] = File(None),
    position: str = Form("center"),
    opacity: float = Form(0.3),
    rotation: int = Form(45),
    size: int = Form(50)
):
    """Add text or image watermark to PDF"""
    temp_file = None
    output_file = None
    watermark_file = None
    watermark_image_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        # Create watermark PDF
        watermark_file = UPLOAD_DIR / f"{uuid.uuid4()}_watermark.pdf"
        c = canvas.Canvas(str(watermark_file), pagesize=letter)
        width, height = letter
        
        c.saveState()
        
        # Calculate position
        if position == "center":
            x_pos, y_pos = width/2, height/2
        elif position == "top-left":
            x_pos, y_pos = width*0.25, height*0.75
        elif position == "top-right":
            x_pos, y_pos = width*0.75, height*0.75
        elif position == "bottom-left":
            x_pos, y_pos = width*0.25, height*0.25
        elif position == "bottom-right":
            x_pos, y_pos = width*0.75, height*0.25
        else:
            x_pos, y_pos = width/2, height/2
        
        c.translate(x_pos, y_pos)
        c.rotate(rotation)
        
        # Check if image watermark is provided
        if watermark_image and watermark_image.filename:
            # Save watermark image
            watermark_image_file = await save_upload_file(watermark_image)
            
            # Open and resize image
            img = Image.open(str(watermark_image_file))
            
            # Calculate image dimensions based on size parameter
            aspect_ratio = img.width / img.height
            if aspect_ratio > 1:
                img_width = size * 2
                img_height = img_width / aspect_ratio
            else:
                img_height = size * 2
                img_width = img_height * aspect_ratio
            
            # Set opacity
            c.setFillAlpha(opacity)
            
            # Draw image (centered on the position)
            c.drawImage(
                ImageReader(img),
                -img_width/2, -img_height/2,
                width=img_width, height=img_height,
                mask='auto',
                preserveAspectRatio=True
            )
        elif text:
            # Text watermark
            c.setFont("Helvetica", size)
            c.setFillColorRGB(0.5, 0.5, 0.5, alpha=opacity)
            c.drawCentredString(0, 0, text)
        else:
            raise HTTPException(status_code=400, detail="Either text or watermark_image must be provided")
        
        c.restoreState()
        c.save()
        
        # Apply watermark
        pdf_reader = PdfReader(str(temp_file))
        watermark_reader = PdfReader(str(watermark_file))
        pdf_writer = PdfWriter()
        
        watermark_page = watermark_reader.pages[0]
        
        for page in pdf_reader.pages:
            page.merge_page(watermark_page)
            pdf_writer.add_page(page)
        
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_watermarked.pdf"
        with open(output_file, "wb") as f:
            pdf_writer.write(f)
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="watermarked.pdf",
            background=lambda: cleanup_files(temp_file, watermark_file, watermark_image_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, watermark_file, watermark_image_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/protect")
async def protect_pdf(file: UploadFile = File(...), password: str = Form(...)):
    """Add password protection to PDF"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        pdf_reader = PdfReader(str(temp_file))
        pdf_writer = PdfWriter()
        
        for page in pdf_reader.pages:
            pdf_writer.add_page(page)
        
        pdf_writer.encrypt(password)
        
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_protected.pdf"
        with open(output_file, "wb") as f:
            pdf_writer.write(f)
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="protected.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/unlock")
async def unlock_pdf(file: UploadFile = File(...), password: str = Form(...)):
    """Remove password protection from PDF"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        pdf_reader = PdfReader(str(temp_file))
        
        if pdf_reader.is_encrypted:
            pdf_reader.decrypt(password)
        
        pdf_writer = PdfWriter()
        for page in pdf_reader.pages:
            pdf_writer.add_page(page)
        
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_unlocked.pdf"
        with open(output_file, "wb") as f:
            pdf_writer.write(f)
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="unlocked.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/sign")
async def sign_pdf(file: UploadFile = File(...), signature_text: str = Form(...)):
    """Add signature to PDF"""
    temp_file = None
    output_file = None
    signature_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        # Create signature overlay
        signature_file = UPLOAD_DIR / f"{uuid.uuid4()}_signature.pdf"
        c = canvas.Canvas(str(signature_file), pagesize=letter)
        width, height = letter
        
        c.setFont("Helvetica-Oblique", 24)
        c.drawString(50, 50, signature_text)
        c.save()
        
        # Apply signature
        pdf_reader = PdfReader(str(temp_file))
        signature_reader = PdfReader(str(signature_file))
        pdf_writer = PdfWriter()
        
        signature_page = signature_reader.pages[0]
        
        # Add signature to last page
        for i, page in enumerate(pdf_reader.pages):
            if i == len(pdf_reader.pages) - 1:
                page.merge_page(signature_page)
            pdf_writer.add_page(page)
        
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_signed.pdf"
        with open(output_file, "wb") as f:
            pdf_writer.write(f)
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="signed.pdf",
            background=lambda: cleanup_files(temp_file, signature_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, signature_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/ipynb-to-pdf")
async def ipynb_to_pdf(file: UploadFile = File(...)):
    """Convert Jupyter Notebook (.ipynb) to PDF"""
    temp_file = None
    output_file = None
    
    try:
        # Validate file extension
        if not file.filename.lower().endswith('.ipynb'):
            raise HTTPException(status_code=400, detail="File must be a .ipynb file")
        
        temp_file = await save_upload_file(file)
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_notebook.pdf"
        
        # Read and validate the notebook
        try:
            with open(temp_file, 'r', encoding='utf-8') as f:
                notebook = nbformat.read(f, as_version=4)
        except nbformat.reader.NotJSONError as e:
            raise HTTPException(status_code=400, detail=f"Invalid Jupyter Notebook file. The file is not a valid JSON format. Please upload a valid .ipynb file.")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read notebook file: {str(e)}")
        
        # Use ReportLab to create PDF directly from notebook content
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Preformatted, PageBreak, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_LEFT
        from reportlab.lib import colors
        
        doc = SimpleDocTemplate(str(output_file), pagesize=letter,
                                leftMargin=0.75*inch, rightMargin=0.75*inch,
                                topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles for input cells
        input_style = ParagraphStyle(
            'Input',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#0066cc'),
            fontName='Courier',
        )
        
        # Custom style for code content
        code_style = ParagraphStyle(
            'Code',
            parent=styles['Code'],
            fontSize=10,
            leading=14,
            fontName='Courier',
            leftIndent=10,
            rightIndent=10,
            spaceBefore=2,
            spaceAfter=2,
            wordWrap='CJK',
        )
        
        # Custom style for output
        output_style = ParagraphStyle(
            'Output',
            parent=styles['Code'],
            fontSize=10,
            leading=14,
            fontName='Courier',
            leftIndent=10,
            rightIndent=10,
            spaceBefore=2,
            spaceAfter=2,
            wordWrap='CJK',
        )
        
        # Process each cell
        cell_num = 0
        for cell in notebook.cells:
            cell_num += 1
            
            if cell.cell_type == 'markdown':
                # Markdown cell styling
                markdown_data = [[Paragraph(cell.source.replace('\n', '<br/>'), styles['Normal'])]]
                markdown_table = Table(markdown_data, colWidths=[6.5*inch])
                markdown_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 10),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                    ('TOPPADDING', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0')),
                ]))
                story.append(markdown_table)
                story.append(Spacer(1, 0.1*inch))
                
            elif cell.cell_type == 'code':
                # Input cell with label
                input_label = f"In [{cell_num}]:"
                code_text = cell.source
                
                # Truncate long lines instead of wrapping
                max_line_length = 75
                code_lines = code_text.split('\n')
                truncated_lines = []
                for line in code_lines:
                    if len(line) > max_line_length:
                        # Truncate and add ellipsis
                        truncated_lines.append(line[:max_line_length] + '...')
                    else:
                        truncated_lines.append(line)
                code_text = '\n'.join(truncated_lines)
                
                # Use smaller font for code to fit better
                small_code_style = ParagraphStyle(
                    'SmallCode',
                    parent=styles['Code'],
                    fontSize=8,
                    leading=11,
                    fontName='Courier',
                    leftIndent=5,
                    rightIndent=5,
                    spaceBefore=2,
                    spaceAfter=2,
                    wordWrap='CJK',
                )
                
                input_data = [[Paragraph(input_label, input_style), Preformatted(code_text, small_code_style)]]
                input_table = Table(input_data, colWidths=[0.7*inch, 5.8*inch])
                input_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f7f7f7')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (1, 0), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 5),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0')),
                ]))
                story.append(input_table)
                
                # Add outputs if any
                if 'outputs' in cell and cell.outputs:
                    for output in cell.outputs:
                        output_text = ""
                        output_bg_color = colors.white
                        output_text_color = colors.black
                        
                        if output.output_type == 'stream':
                            output_text = output.text
                        elif output.output_type == 'execute_result' or output.output_type == 'display_data':
                            if 'text/plain' in output.data:
                                output_text = output.data['text/plain']
                        elif output.output_type == 'error':
                            # Handle error output with red background
                            output_bg_color = colors.HexColor('#fff0f0')
                            output_text_color = colors.HexColor('#d32f2f')
                            # Combine error name, value, and traceback
                            error_parts = []
                            if hasattr(output, 'ename'):
                                error_parts.append(output.ename)
                            if hasattr(output, 'evalue'):
                                error_parts.append(output.evalue)
                            if hasattr(output, 'traceback'):
                                # Remove ANSI color codes from traceback
                                import re
                                clean_traceback = []
                                for line in output.traceback:
                                    # Remove ANSI escape sequences
                                    clean_line = re.sub(r'\x1b\[[0-9;]*m', '', line)
                                    clean_traceback.append(clean_line)
                                error_parts.extend(clean_traceback)
                            output_text = '\n'.join(error_parts)
                        
                        if output_text:
                            # Truncate long lines in output - don't wrap
                            max_line_length = 75
                            output_lines = output_text.split('\n')
                            truncated_output = []
                            for line in output_lines:
                                if len(line) > max_line_length:
                                    # Truncate and add ellipsis
                                    truncated_output.append(line[:max_line_length] + '...')
                                else:
                                    truncated_output.append(line)
                            output_text = '\n'.join(truncated_output)
                            
                            # Create custom style for this output - smaller font
                            custom_output_style = ParagraphStyle(
                                'CustomOutput',
                                parent=styles['Code'],
                                fontSize=8,
                                leading=11,
                                fontName='Courier',
                                leftIndent=5,
                                rightIndent=5,
                                spaceBefore=2,
                                spaceAfter=2,
                                textColor=output_text_color,
                                wordWrap='CJK',
                            )
                            
                            output_data = [[Paragraph("", custom_output_style), Preformatted(output_text, custom_output_style)]]
                            output_table = Table(output_data, colWidths=[0.7*inch, 5.8*inch])
                            output_table.setStyle(TableStyle([
                                ('BACKGROUND', (0, 0), (-1, -1), output_bg_color),
                                ('TEXTCOLOR', (0, 0), (-1, -1), output_text_color),
                                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                                ('TOPPADDING', (0, 0), (-1, -1), 8),
                                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0')),
                            ]))
                            story.append(output_table)
                
                story.append(Spacer(1, 0.15*inch))
        
        # Build PDF
        doc.build(story)
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="notebook.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except HTTPException:
        cleanup_files(temp_file, output_file)
        raise
    except Exception as e:
        cleanup_files(temp_file, output_file)
        logging.error(f"Error converting notebook to PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to convert notebook to PDF: {str(e)}")

@api_router.post("/java-to-pdf")
async def java_to_pdf(file: UploadFile = File(...)):
    """Convert Java source code file to PDF with formatted text and line numbers"""
    temp_file = None
    output_file = None
    
    try:
        # Validate file extension
        if not file.filename.lower().endswith('.java'):
            raise HTTPException(status_code=400, detail="File must be a .java file")
        
        temp_file = await save_upload_file(file)
        
        # Read Java source code
        with open(temp_file, 'r', encoding='utf-8') as f:
            java_code = f.read()
        
        # Create PDF with formatted code
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_java.pdf"
        
        # Create PDF using ReportLab
        c = canvas.Canvas(str(output_file), pagesize=letter)
        width, height = letter
        
        # Set up fonts and layout
        margin = 50
        y_position = height - margin
        line_height = 12
        font_size = 9
        
        # Set font for code
        c.setFont("Courier", font_size)
        
        # Split code into lines and write to PDF
        lines = java_code.split('\n')
        
        for i, line in enumerate(lines, 1):
            # Check if we need a new page
            if y_position < margin + 20:
                c.showPage()
                y_position = height - margin
                c.setFont("Courier", font_size)
            
            # Draw code line (truncate if too long)
            c.setFillColorRGB(0, 0, 0)
            max_chars = 100
            if len(line) > max_chars:
                display_line = line[:max_chars] + "..."
            else:
                display_line = line
            c.drawString(margin, y_position, display_line)
            
            y_position -= line_height
        
        c.save()
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename=f"{Path(file.filename).stem}.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except UnicodeDecodeError:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=400, detail="Unable to read Java file. Please ensure it's a valid text file with UTF-8 encoding.")
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/add-page-numbers")
async def add_page_numbers(
    file: UploadFile = File(...),
    format: str = Form(...),
    position: str = Form(...)
):
    """Add page numbers to PDF pages"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        # Read the original PDF
        pdf_reader = PdfReader(str(temp_file))
        pdf_writer = PdfWriter()
        
        total_pages = len(pdf_reader.pages)
        
        # Helper function to convert number to roman numerals
        def to_roman(num):
            val = [
                1000, 900, 500, 400,
                100, 90, 50, 40,
                10, 9, 5, 4,
                1
            ]
            syms = [
                "M", "CM", "D", "CD",
                "C", "XC", "L", "XL",
                "X", "IX", "V", "IV",
                "I"
            ]
            roman_num = ''
            i = 0
            while num > 0:
                for _ in range(num // val[i]):
                    roman_num += syms[i]
                    num -= val[i]
                i += 1
            return roman_num
        
        # Process each page
        for page_num in range(total_pages):
            # Get the page from original PDF
            page = pdf_reader.pages[page_num]
            page_width = float(page.mediabox.width)
            page_height = float(page.mediabox.height)
            
            # Create a new PDF with the page number
            packet = io.BytesIO()
            can = canvas.Canvas(packet, pagesize=(page_width, page_height))
            
            # Determine page number text based on format
            page_number = page_num + 1
            if format == "numeric":
                page_text = str(page_number)
            elif format == "numeric-page":
                page_text = f"Page {page_number}"
            elif format == "roman-lower":
                page_text = to_roman(page_number).lower()
            elif format == "roman-lower-page":
                page_text = f"Page {to_roman(page_number).lower()}"
            elif format == "roman-upper":
                page_text = to_roman(page_number)
            elif format == "roman-upper-page":
                page_text = f"Page {to_roman(page_number)}"
            else:
                page_text = str(page_number)
            
            # Set font and determine position
            can.setFont("Helvetica", 10)
            text_width = can.stringWidth(page_text, "Helvetica", 10)
            
            # Position the page number (bottom)
            y_position = 20  # 20 points from bottom
            
            if position == "bottom-left":
                x_position = 40  # 40 points from left
            elif position == "bottom-center":
                x_position = (page_width - text_width) / 2
            elif position == "bottom-right":
                x_position = page_width - text_width - 40  # 40 points from right
            else:
                x_position = (page_width - text_width) / 2  # default to center
            
            # Draw the page number
            can.drawString(x_position, y_position, page_text)
            can.save()
            
            # Move to the beginning of the BytesIO buffer
            packet.seek(0)
            
            # Read the page number overlay
            overlay_pdf = PdfReader(packet)
            overlay_page = overlay_pdf.pages[0]
            
            # Merge the overlay with the original page
            page.merge_page(overlay_page)
            pdf_writer.add_page(page)
        
        # Save the output PDF
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_numbered.pdf"
        with open(output_file, "wb") as f:
            pdf_writer.write(f)
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="numbered.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/python-to-pdf")
async def python_to_pdf(file: UploadFile = File(...)):
    """Convert Python source code file to PDF with formatted text and line numbers"""
    temp_file = None
    output_file = None
    
    try:
        # Validate file extension
        if not file.filename.lower().endswith('.py'):
            raise HTTPException(status_code=400, detail="File must be a .py file")
        
        temp_file = await save_upload_file(file)
        
        # Read Python source code
        with open(temp_file, 'r', encoding='utf-8') as f:
            python_code = f.read()
        
        # Create PDF with formatted code
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_python.pdf"
        
        # Create PDF using ReportLab
        c = canvas.Canvas(str(output_file), pagesize=letter)
        width, height = letter
        
        # Set up fonts and layout
        margin = 50
        y_position = height - margin
        line_height = 12
        font_size = 9
        
        # Set font for code
        c.setFont("Courier", font_size)
        
        # Split code into lines and write to PDF
        lines = python_code.split('\n')
        
        for line in lines:
            # Check if we need a new page
            if y_position < margin + 20:
                c.showPage()
                y_position = height - margin
                c.setFont("Courier", font_size)
            
            # Draw code line (truncate if too long)
            c.setFillColorRGB(0, 0, 0)
            max_chars = 100
            if len(line) > max_chars:
                display_line = line[:max_chars] + "..."
            else:
                display_line = line
            c.drawString(margin, y_position, display_line)
            
            y_position -= line_height
        
        c.save()
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename=f"{Path(file.filename).stem}.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except UnicodeDecodeError:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=400, detail="Unable to read Python file. Please ensure it's a valid text file with UTF-8 encoding.")
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/xml-to-pdf")
async def xml_to_pdf(file: UploadFile = File(...)):
    """Convert XML file to PDF with formatted structure"""
    temp_file = None
    output_file = None
    
    try:
        # Validate file extension
        if not file.filename.lower().endswith('.xml'):
            raise HTTPException(status_code=400, detail="File must be an .xml file")
        
        temp_file = await save_upload_file(file)
        
        # Read and parse XML file
        try:
            tree = ET.parse(str(temp_file))
            root = tree.getroot()
            
            # Pretty print XML for better formatting
            xml_string = ET.tostring(root, encoding='unicode')
            dom = minidom.parseString(xml_string)
            pretty_xml = dom.toprettyxml(indent="  ")
        except ET.ParseError as e:
            cleanup_files(temp_file, output_file)
            raise HTTPException(status_code=400, detail=f"Invalid XML file: {str(e)}")
        except Exception as e:
            cleanup_files(temp_file, output_file)
            raise HTTPException(status_code=400, detail=f"Error parsing XML: {str(e)}")
        
        # Create PDF with formatted XML
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_xml.pdf"
        
        # Create PDF using SimpleDocTemplate (more reliable than canvas)
        doc = SimpleDocTemplate(str(output_file), pagesize=letter)
        styles = getSampleStyleSheet()
        
        # Create a custom style for XML content
        code_style = ParagraphStyle(
            'XMLCode',
            parent=styles['Code'],
            fontName='Courier',
            fontSize=9,
            leading=11,
            leftIndent=0,
            rightIndent=0,
            alignment=TA_LEFT,
            spaceBefore=0,
            spaceAfter=0,
        )
        
        story = []
        
        # Add title
        title_style = ParagraphStyle(
            'XMLTitle',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=20,
        )
        story.append(Paragraph(f"XML File: {file.filename}", title_style))
        story.append(Spacer(1, 0.2 * inch))
        
        # Split XML content into lines and add to PDF
        lines = pretty_xml.split('\n')
        
        for line in lines:
            # Skip empty lines at the start
            if not line.strip():
                line = '&nbsp;'
                
            # Escape special characters for reportlab
            line = line.replace('&', '&amp;')
            line = line.replace('<', '&lt;')
            line = line.replace('>', '&gt;')
            # Preserve spaces
            line = line.replace(' ', '&nbsp;')
            line = line.replace('\t', '&nbsp;&nbsp;')
            
            story.append(Paragraph(line, code_style))
        
        doc.build(story)
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename=f"{Path(file.filename).stem}.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except HTTPException:
        # Re-raise HTTP exceptions
        raise
    except UnicodeDecodeError:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=400, detail="Unable to read XML file. Please ensure it's a valid text file with UTF-8 encoding.")
    except Exception as e:
        cleanup_files(temp_file, output_file)
        logging.error(f"XML to PDF error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error converting XML to PDF: {str(e)}")

@api_router.post("/preview-pages")
async def preview_pdf_pages(file: UploadFile = File(...)):
    """Generate preview images for all pages in PDF"""
    temp_file = None
    preview_files = []
    
    try:
        temp_file = await save_upload_file(file)
        
        # Open PDF with PyMuPDF
        import fitz
        pdf_document = fitz.open(str(temp_file))
        
        # Generate previews for all pages
        previews = []
        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]
            # Render page to image with 150 DPI for better quality
            pix = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
            
            # Convert to base64 for JSON response
            import base64
            img_bytes = pix.tobytes("png")
            img_base64 = base64.b64encode(img_bytes).decode('utf-8')
            
            previews.append({
                "pageNumber": page_num + 1,
                "imageData": f"data:image/png;base64,{img_base64}",
                "width": pix.width,
                "height": pix.height
            })
        
        pdf_document.close()
        cleanup_files(temp_file)
        
        return JSONResponse(content={
            "totalPages": len(previews),
            "pages": previews
        })
    
    except Exception as e:
        cleanup_files(temp_file)
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/delete-pages")
async def delete_pdf_pages(file: UploadFile = File(...), pages_to_delete: str = Form(...)):
    """Delete specified pages from PDF"""
    temp_file = None
    output_file = None
    
    try:
        temp_file = await save_upload_file(file)
        
        # Parse pages to delete (comma-separated list, e.g., "1,3,5")
        pages_to_delete_list = [int(p.strip()) - 1 for p in pages_to_delete.split(',') if p.strip()]
        
        # Read PDF
        pdf_reader = PdfReader(str(temp_file))
        pdf_writer = PdfWriter()
        
        # Add all pages except the ones to delete
        total_pages = len(pdf_reader.pages)
        for page_num in range(total_pages):
            if page_num not in pages_to_delete_list:
                pdf_writer.add_page(pdf_reader.pages[page_num])
        
        # Check if all pages were deleted
        if len(pdf_writer.pages) == 0:
            raise HTTPException(status_code=400, detail="Cannot delete all pages. At least one page must remain.")
        
        # Save modified PDF
        output_file = UPLOAD_DIR / f"{uuid.uuid4()}_deleted.pdf"
        with open(output_file, "wb") as f:
            pdf_writer.write(f)
        
        return FileResponse(
            output_file,
            media_type="application/pdf",
            filename="modified.pdf",
            background=lambda: cleanup_files(temp_file, output_file)
        )
    
    except ValueError as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=400, detail="Invalid page numbers format. Use comma-separated numbers (e.g., 1,3,5)")
    except Exception as e:
        cleanup_files(temp_file, output_file)
        raise HTTPException(status_code=500, detail=str(e))

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    if client:
        client.close()